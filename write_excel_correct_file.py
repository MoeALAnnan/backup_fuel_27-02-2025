import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
import os
from write2 import main
import pickle

def main_excel():
    with open('dictionary_page0.pkl', 'rb') as file:
                dictionary_page0 = pickle.load(file)
    print(dictionary_page0)
    def apply_border_to_merged_cells(sheet, start_row, end_row, start_col, end_col):
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                cell.border = border
                cell.alignment = alignment
    base_directory = r'C:\Users\Momo\Desktop'
    # Define the absolute file paths
    excel_file = os.path.join(base_directory, 'test_1.xlsx')
    

    # Check if the Excel file exists
    if not os.path.exists(excel_file):
        # Create a new workbook if it doesn't exist
        wb = Workbook()
        sheet = wb.active

        # Merge cells B8 to I8 and J8 to L8
        sheet.merge_cells('B8:I8')
        sheet.merge_cells('J8:L8')
        sheet.merge_cells('M8:P8')
        sheet.merge_cells('F10:G10')
        sheet.merge_cells('I10:J10')
        sheet.merge_cells('N10:O10')
        sheet.merge_cells('B12:C12')
        sheet.merge_cells('C14:E14')
        sheet.merge_cells('F14:H14')
        sheet.merge_cells('I14:N14')
        sheet.merge_cells('B22:B23')
        sheet.merge_cells('C22:D22')
        sheet.merge_cells('E22:F22')
        sheet.merge_cells('G22:G23')
        sheet.merge_cells('B24:B25')
        sheet.merge_cells('C24:C25')
        sheet.merge_cells('D24:D25')
        sheet.merge_cells('E24:E25')
        sheet.merge_cells('F24:F25')
        sheet.merge_cells('G24:G25')
        sheet.merge_cells('B26:B27')
        sheet.merge_cells('C26:C27')
        sheet.merge_cells('D26:D27')
        sheet.merge_cells('E26:E27')
        sheet.merge_cells('F26:F27')
        sheet.merge_cells('G26:G27')

        # Set the row height to make sure both merged cells are visible
        sheet.row_dimensions[8].height = 30
        sheet.row_dimensions[15].height = 30

        # Apply borders to merged cells
        apply_border_to_merged_cells(sheet, 8, 8, 2, 16)  # Apply borders to B8:I8
        apply_border_to_merged_cells(sheet, 10, 10, 2, 16)  # Apply borders to M8:P

        apply_border_to_merged_cells(sheet, 12, 12, 2, 16)  # Apply borders to M8:P
        apply_border_to_merged_cells(sheet, 14, 19, 2, 16)  # Apply borders to M8:P
        apply_border_to_merged_cells(sheet, 22, 27, 2, 7)  # Apply borders to M8:P
        apply_border_to_merged_cells(sheet, 22, 24, 10, 11)  # Apply borders to M8:P
        apply_border_to_merged_cells(sheet, 26, 27, 10, 11)  # Apply borders to M8:P
        apply_border_to_merged_cells(sheet, 29, 30, 10, 11)  # Apply borders to M8:P
        # Set the values in the merged cells
        sheet['B8'] = 'GTI : 8.28.165 FUEL CALIBRATION.test'
        sheet['J8'] = dictionary_page0['KILO/LIVRE']
        sheet['M8'] = dictionary_page0['Engine Type']
        sheet['B10'] = 'A/C'
        sheet['C10'] = dictionary_page0['MSN']
        sheet['D10'] = 'VERSION'
        sheet['E10'] = dictionary_page0['VERSION']
        sheet['F10'] = 'FQI P/N'
        sheet['H10'] = dictionary_page0['Part Number']
        sheet['I10'] = 'FQI S/N'
        sheet['K10'] = dictionary_page0['Serial Number']
        sheet['L10'] = 'stamp'
        sheet['M10'] = dictionary_page0['Stamp']
        sheet['N10'] = 'Date'
        sheet['P10'] = dictionary_page0['Date']
        sheet['B12'] = 'Pression Pompes (bar)'
        sheet['D12'] = 'L.T.1'
        sheet['E12'] = dictionary_page0['Left Pump 1 P(bar)']
        sheet['F12'] = 'L.T.2'
        sheet['G12'] = dictionary_page0['Left Pump 2 P(bar)']
        sheet['H12'] = 'R.T.1'
        sheet['I12'] = dictionary_page0['Right Pump 1 P(bar)']
        sheet['J12'] = 'R.T.2'
        sheet['K12'] = dictionary_page0['Right Pump 2 P(bar)']
        sheet['B14'] = 'LEVELS'
        sheet['C14'] = 'UPLIFTED FUEL PARAMETERS'
        sheet['F14'] = 'INDICATOR'
        sheet['I14'] = 'ACT FITTED (si option)'
        sheet['C15'] = 'VOLUME (V)'
        sheet['C17'] = dictionary_page0['stp1 volume']
        sheet['C18'] = dictionary_page0['stp2 volume']
        sheet['C19'] = dictionary_page0['stp3 volume']
        sheet['D15'] = 'Density (p)'
        sheet['D17'] = dictionary_page0['Fuel_Density']
        sheet['D18'] = dictionary_page0['Fuel_Density']
        sheet['D19'] = dictionary_page0['Fuel_Density']
        sheet['K22'] = dictionary_page0['Fuel_Density']
        sheet['E15'] = 'Uplifted mass (m)'
        sheet['E17'] = dictionary_page0['step1 uplifted']
        sheet['E18'] = dictionary_page0['step2 uplifted']
        sheet['E19'] = dictionary_page0['step3 uplifted']
        sheet['F15'] = 'L.T'
        sheet['F17'] = dictionary_page0['step1 left']
        sheet['F18'] = dictionary_page0['step2 left']
        sheet['F19'] = dictionary_page0['step3 left']
        sheet['G15'] = 'CTR.T'
        sheet['G17'] = '0'
        sheet['G18'] = dictionary_page0['step2 center']
        sheet['G19'] = dictionary_page0['step3 center']
        sheet['H15'] = 'R.T'
        sheet['H17'] = dictionary_page0['step1 right']
        sheet['H18'] = dictionary_page0['step2 right']
        sheet['H19'] = dictionary_page0['step3 right']
        sheet['I15'] = '*RCT'
        sheet['J15'] = '*AFT 1'
        sheet['K15'] = '*AFT 2'
        sheet['L15'] = '*AFT 3'
        sheet['M15'] = '*FWD'
        sheet['N15'] = '*AFT 5'
        sheet['O15'] = '*TOTAL FOB'
        sheet['P15'] = '*FQI accuracy (0)'
        sheet['P19'] = dictionary_page0['Accuracy %']
        sheet['B16'] = '0'
        sheet['B17'] = '4 (STEP1)'
        sheet['B18'] = 'CTR HL'
        sheet['B19'] = 'HI LVL (STEP3)'
        sheet['C22'] = 'ECAM (unit)'
        sheet['E22'] = 'Calcul (unit/min)'
        sheet['C23'] = 'Start'
        sheet['C24'] = dictionary_page0['Jet Pump 1 start mass']
        sheet['C26'] = dictionary_page0['Jet Pump 2 start mass']
        sheet['D23'] = 'Stop(5 min)'
        sheet['D24'] = dictionary_page0['Jet Pump 1 end mass']
        sheet['D26'] = dictionary_page0['Jet Pump 2 end mass']
        sheet['E23'] = 'Rate'
        sheet['E24'] = dictionary_page0['Mass Rate 1 (kg/min)']
        sheet['E26'] = dictionary_page0['Mass Rate 2 (kg/min)']
        sheet['F23'] = 'Limit'
        sheet['G22'] = 'Tolerances'
        sheet['B24'] = 'Left'
        sheet['B26'] = 'Right'
        sheet['J22'] = 'Density (p)'
        sheet['J23'] = '(t°c) FUEL'
        sheet['K23'] = dictionary_page0['Fuel_Temperature']
        sheet['J24'] = '(t°c) AMB'
        sheet['K24'] = dictionary_page0['Ambient_Temperature']
        sheet['J26'] = 'PITCH'
        sheet['K26'] = dictionary_page0['Pitch']
        sheet['J27'] = 'ROLL'
        sheet['K27'] = dictionary_page0['Roll']
        sheet['J29'] = 'L.T'
        sheet['J30'] = 'R.T'
    
        # Save the workbook
        wb.save(excel_file)
        # this part will be used to update the existing excel sheet 
    else:
        # Open the existing workbook
        wb = xw.Book(excel_file)

        # Define the sheet after opening the workbook
        sheet = wb.sheets[0]

        # Update the values in the merged cells
        sheet.range('B8').value = 'GTI : 8.28.165 FUEL CALIBRATION.test2'
        sheet.range('J8').value = 'UNIT2'
        sheet.range('M8').value = 'Engine Type'

        # Save the workbook
        wb.save(excel_file)
        wb.close()

    # Create PDF from Excel using built-in functionality
    
    wb = xw.Book(excel_file)

    # Activate the appropriate sheet before exporting
    sheet = wb.sheets[0]
    sheet.activate()

    os.startfile(excel_file)
    main()
